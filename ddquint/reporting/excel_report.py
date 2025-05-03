"""
Robust Excel report generation module for ddQuint
Creates a detailed Excel report with copy number results in a specific grid layout
with proper borders and merged headers
"""

import os
import numpy as np
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Define plate layout constants
ROW_LABELS = list('ABCDEFGH')
COL_LABELS = [str(i) for i in range(1, 13)]  # 1-12

# Define cell colors
NORMAL_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
OUTLIER_FILL = PatternFill(start_color="E6B8E6", end_color="E6B8E6", fill_type="solid")  # Light purple
OUTLIER_VALUE_FILL = PatternFill(start_color="D070D0", end_color="D070D0", fill_type="solid")  # Darker purple

# Define border styles - revised for better clarity
thin = Side(style='thin')
thick = Side(style='thick')

# Standard borders
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)

# Specialized borders for well boundaries
# For top borders
top_thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
top_thick_border = Border(left=thin, right=thin, top=thick, bottom=thin)
top_thick_left_border = Border(left=thick, right=thin, top=thick, bottom=thin)
top_thick_right_border = Border(left=thin, right=thick, top=thick, bottom=thin)

# For bottom borders
bottom_thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
bottom_thick_border = Border(left=thin, right=thin, top=thin, bottom=thick)
bottom_thick_left_border = Border(left=thick, right=thin, top=thin, bottom=thick)
bottom_thick_right_border = Border(left=thin, right=thick, top=thin, bottom=thick)

# For left/right borders
left_thick_border = Border(left=thick, right=thin, top=thin, bottom=thin)
right_thick_border = Border(left=thin, right=thick, top=thin, bottom=thin)

# Corner borders - completely redefined for clarity
top_left_corner = Border(left=thick, right=thin, top=thick, bottom=thin)
top_right_corner = Border(left=thin, right=thick, top=thick, bottom=thin)
bottom_left_corner = Border(left=thick, right=thin, top=thin, bottom=thick)
bottom_right_corner = Border(left=thin, right=thick, top=thin, bottom=thick)

def create_excel_report(results, output_path, template_path=None):
    """
    Create an Excel report with all analysis results.
    
    Args:
        results (list): List of result dictionaries for each well
        output_path (str): Path to save the Excel report
        template_path (str, optional): Not used, kept for API compatibility
        
    Returns:
        str: Path to the saved Excel report
    """
    print(f"Creating Excel report...")
    
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Plate Results"

        # Create the full grid layout without any merging first
        create_grid_layout_without_merging(ws, results)
        
        # Apply all cell merges AFTER setting values
        apply_cell_merges(ws)
        
        # Apply formatting and borders
        apply_formatting(ws)
        
        # Save the workbook
        try:
            wb.save(output_path)
            return output_path
        except Exception as e:
            print(f"Error saving Excel report: {str(e)}")
            traceback.print_exc()
            # Try to save with a different name if there was an error
            try:
                alt_path = os.path.join(os.path.dirname(output_path), "Plate_Results_alt.xlsx")
                wb.save(alt_path)
                print(f"Saved alternative report to: {alt_path}")
                return alt_path
            except Exception as e2:
                print(f"Error saving alternative report: {str(e2)}")
                traceback.print_exc()
                return None
    except Exception as e:
        print(f"Error creating Excel report: {str(e)}")
        traceback.print_exc()
        return None

def create_grid_layout_without_merging(ws, results):
    """
    Create the 12x8 grid layout for the 96-well plate without any cell merging.
    
    Args:
        ws: Worksheet to modify
        results: List of result dictionaries
    """
    # Set header row values
    setup_header_values(ws)
    
    # Convert results list to a dictionary keyed by well ID for easy lookup
    result_dict = {r.get('well', ''): r for r in results if r.get('well') is not None}
    
    # Process each row (A-H)
    for row_idx, row_label in enumerate(ROW_LABELS):
        # Starting row for this plate row (each well takes 6 rows)
        start_row = row_idx * 6 + 3  # Start from row 3 (after headers)
        
        # Add row label in column A
        row_label_cell = ws.cell(row=start_row, column=1)
        row_label_cell.value = row_label
        row_label_cell.font = Font(bold=True)
        row_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Process each column (1-12)
        for col_idx, col_label in enumerate(COL_LABELS):
            well_id = f"{row_label}{col_label.zfill(2)}"
            
            # Starting column for this well (each well takes 3 columns)
            start_col = col_idx * 3 + 2  # Start from column B (column 2)
            
            # Fill in data for this well without merging
            add_well_data(ws, start_row, start_col, well_id, result_dict.get(well_id))

def setup_header_values(ws):
    """
    Set up the header row values without merging.
    """
    # Row 1: Column numbers (1-12)
    for col_idx in range(1, 13):
        # Each well takes 3 columns
        cell_col = (col_idx - 1) * 3 + 2  # Start at column B (column 2)
        
        # Set column number in the first column of the well
        cell = ws.cell(row=1, column=cell_col)
        cell.value = col_idx
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
        # The rest of the columns in this merged range will be merged later
    
    # Row 2: "abs." and "rel." labels
    for col_idx in range(1, 13):
        base_col = (col_idx - 1) * 3 + 2
        
        # First column is blank (for chromosome name)
        first_cell = ws.cell(row=2, column=base_col)
        first_cell.value = ""
        
        # Second column: "abs."
        abs_cell = ws.cell(row=2, column=base_col + 1)
        abs_cell.value = "abs."
        abs_cell.font = Font(size=9)
        abs_cell.alignment = Alignment(horizontal='center')
        
        # Third column: "rel."
        rel_cell = ws.cell(row=2, column=base_col + 2)
        rel_cell.value = "rel."
        rel_cell.font = Font(size=9)
        rel_cell.alignment = Alignment(horizontal='center')

def add_well_data(ws, start_row, start_col, well_id, result):
    """
    Add data for a single well without merging.
    
    Args:
        ws: Worksheet to modify
        start_row: Starting row for this well
        start_col: Starting column for this well
        well_id: Well identifier (e.g., 'A1')
        result: Result dictionary for this well, or None if no data
    """
    # Add "name" header for the well (will be merged later)
    name_cell = ws.cell(row=start_row, column=start_col)
    name_cell.value = "name"
    name_cell.alignment = Alignment(horizontal='center')
    
    # Will set blank values to cells that will be merged
    ws.cell(row=start_row, column=start_col+1).value = ""
    ws.cell(row=start_row, column=start_col+2).value = ""
    
    # If we have a result for this well, fill in the data
    if result:
        # Get the sample name from the filename
        filename = result.get('filename', '')
        sample_name = os.path.splitext(filename)[0] if filename else well_id
        
        # Add sample name to the "name" row (will be merged later)
        sample_cell = ws.cell(row=start_row+1, column=start_col)
        sample_cell.value = sample_name
        
        # These will be merged later but set to empty now
        ws.cell(row=start_row+1, column=start_col+1).value = ""
        ws.cell(row=start_row+1, column=start_col+2).value = ""
        
        # Get count data and copy numbers
        counts = result.get('counts', {})
        copy_numbers = result.get('copy_numbers', {})
        has_outlier = result.get('has_outlier', False)
        
        # Create cells for each chromosome
        for chrom_idx, chrom_suffix in enumerate(range(1, 6)):
            chrom = f"Chr{chrom_suffix}"
            chrom_key = f"Chrom{chrom_suffix}"  # Map to keys in the result data
            
            # Row for this chromosome
            chrom_row = start_row + 1 + chrom_idx
            
            # Chromosome name cell
            chrom_cell = ws.cell(row=chrom_row, column=start_col)
            chrom_cell.value = chrom
            
            # Absolute count cell
            abs_cell = ws.cell(row=chrom_row, column=start_col+1)
            abs_count = counts.get(chrom_key, 0)
            abs_cell.value = abs_count if abs_count > 0 else None
            
            # Relative count cell
            rel_cell = ws.cell(row=chrom_row, column=start_col+2)
            rel_count = copy_numbers.get(chrom_key)
            if rel_count is not None:
                rel_cell.value = round(rel_count, 2)
                rel_cell.number_format = '0.00'
            
            # Apply highlighting for outliers
            if has_outlier:
                # Base coloring for all cells in outlier wells
                chrom_cell.fill = OUTLIER_FILL
                abs_cell.fill = OUTLIER_FILL
                rel_cell.fill = OUTLIER_FILL
                
                # Darker coloring for the specific outlier value
                if rel_count is not None and abs(rel_count - 1.0) > 0.15:
                    rel_cell.fill = OUTLIER_VALUE_FILL
    else:
        # If no data for this well, just add the chromosome labels
        for chrom_idx, chrom_suffix in enumerate(range(1, 6)):
            chrom = f"Chr{chrom_suffix}"
            chrom_row = start_row + 1 + chrom_idx
            chrom_cell = ws.cell(row=chrom_row, column=start_col)
            chrom_cell.value = chrom

def apply_cell_merges(ws):
    """
    Apply all cell merges after setting all cell values.
    """
    # Merge cells in header row
    for col_idx in range(1, 13):
        cell_col = (col_idx - 1) * 3 + 2
        ws.merge_cells(start_row=1, start_column=cell_col, end_row=1, end_column=cell_col + 2)
    
    # Merge row labels
    for row_idx in range(len(ROW_LABELS)):
        start_row = row_idx * 6 + 3
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+5, end_column=1)
    
    # Merge well name headers and sample names
    for row_idx in range(len(ROW_LABELS)):
        for col_idx in range(len(COL_LABELS)):
            start_row = row_idx * 6 + 3
            start_col = col_idx * 3 + 2
            
            # Merge the name header
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+2)


def apply_formatting(ws):
    """
    Apply formatting to the entire worksheet with proper borders.
    
    Args:
        ws: Worksheet to format
    """
    # Get maximum row and column in use
    max_row = ws.max_row
    max_col = 37  # 12 wells × 3 columns each + 1 for row labels
    
    # First pass: apply thin borders to all cells
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    
    # Second pass: apply thick borders for plate boundaries
    apply_plate_boundaries(ws)
    
    # Set column widths
    ws.column_dimensions['A'].width = 3  # Row labels
    
    # Each well takes 3 columns
    for i in range(1, 37):  # 12 wells × 3 columns each = 36 columns
        col_letter = get_column_letter(i+1)  # +1 because we start at column B
        if i % 3 == 1:  # First column of each well (chromosome name)
            ws.column_dimensions[col_letter].width = 5
        else:  # Data columns
            ws.column_dimensions[col_letter].width = 6
    
    # Freeze the header rows and first column
    ws.freeze_panes = ws.cell(row=3, column=2)

def apply_plate_boundaries(ws):
    """
    Apply thick borders around each well and at the plate boundaries.
    This function applies consistent thick borders throughout the plate.
    """
    # For each row of wells (A-H)
    for row_idx, row_label in enumerate(ROW_LABELS):
        # Start row for this plate row (each well takes 6 rows)
        start_row = row_idx * 6 + 3
        end_row = start_row + 5  # End row for this plate row
        
        # For each column of wells (1-12)
        for col_idx, col_label in enumerate(COL_LABELS):
            # Start and end columns for this well (each well is 3 columns wide)
            start_col = col_idx * 3 + 2
            end_col = start_col + 2
            
            # Apply appropriate borders to each cell in this well
            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    # Get existing cell and its current border
                    cell = ws.cell(row=r, column=c)
                    
                    # Create a new border with appropriate sides
                    left_side = thick if c == start_col else thin
                    right_side = thick if c == end_col else thin
                    top_side = thick if r == start_row else thin
                    bottom_side = thick if r == end_row else thin
                    
                    # Apply the new border
                    cell.border = Border(
                        left=left_side,
                        right=right_side,
                        top=top_side,
                        bottom=bottom_side
                    )