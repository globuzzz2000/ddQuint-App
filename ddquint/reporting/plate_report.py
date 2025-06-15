#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel report generation module for ddQuint with dynamic chromosome support.

Generates Excel reports in 96-well plate format with copy number data organized
by well position. Supports both standard and rotated layouts with appropriate
highlighting for aneuploidies and buffer zones.
"""

import os
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

from ..config import Config, ReportGenerationError

logger = logging.getLogger(__name__)

# Define plate layout constants
DEFAULT_COL_LABELS = list('ABCDEFGH')  # Vertical labels (columns)
DEFAULT_ROW_LABELS = [str(i) for i in range(1, 13)]  # Horizontal labels (rows)

# Define cell colors
NORMAL_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ANEUPLOIDY_FILL = PatternFill(start_color="E6B8E6", end_color="E6B8E6", fill_type="solid")
ANEUPLOIDY_VALUE_FILL = PatternFill(start_color="D070D0", end_color="D070D0", fill_type="solid")
BUFFER_ZONE_FILL = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
BUFFER_ZONE_VALUE_FILL = PatternFill(start_color="B0B0B0", end_color="B0B0B0", fill_type="solid")

# Define border styles
thin = Side(style='thin')
thick = Side(style='thick')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)


def create_plate_report(results, output_path, template_path=None, rotated=False):
    """
    Create an Excel report with all analysis results in plate format.
    
    Args:
        results (list): List of result dictionaries for each well
        output_path (str): Path to save the Excel report
        template_path (str, optional): Not used, kept for API compatibility
        rotated (bool): If True, use rotated layout (1-12 as rows, A-H as columns)
        
    Returns:
        str: Path to the saved Excel report
        
    Raises:
        ReportGenerationError: If report creation fails
        ValueError: If results data is invalid
    """
    config = Config.get_instance()
    
    if not results:
        error_msg = "No results provided for plate report generation"
        logger.error(error_msg)
        raise ValueError(error_msg)
    
    # Get the number of chromosomes from config
    chromosome_keys = config.get_chromosome_keys()
    num_chromosomes = len(chromosome_keys)
    
    logger.debug(f"Creating Excel report for {len(results)} results with {num_chromosomes} chromosomes")
    logger.debug(f"Output path: {output_path}")
    logger.debug(f"Rotated layout: {rotated}")
    
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Plate Results"
        logger.debug("Created new workbook")

        # Create the full grid layout without any merging first
        create_grid_layout_without_merging(ws, results, num_chromosomes, rotated)
        
        # Apply all cell merges AFTER setting values
        apply_cell_merges(ws, num_chromosomes, rotated)
        
        # Apply formatting and borders
        apply_formatting(ws, num_chromosomes, rotated)
        
        # Save the workbook
        wb.save(output_path)
        logger.debug(f"Excel report saved successfully to {output_path}")
        return output_path
        
    except Exception as e:
        error_msg = f"Error creating Excel report for {os.path.basename(output_path)}: {str(e)}"
        logger.error(error_msg)
        logger.debug(f"Error details: {str(e)}", exc_info=True)
        
        # Try to save with alternative name if there was an error
        try:
            alt_path = os.path.join(os.path.dirname(output_path), "Plate_Results_alt.xlsx")
            wb.save(alt_path)
            logger.info(f"Saved alternative report to: {alt_path}")
            return alt_path
        except Exception as e2:
            logger.error(f"Error saving alternative report: {str(e2)}")
            logger.debug(f"Error details: {str(e2)}", exc_info=True)
            raise ReportGenerationError(error_msg) from e


def create_grid_layout_without_merging(ws, results, num_chromosomes, rotated=False):
    """
    Create the grid layout for the 96-well plate without any cell merging.
    
    Args:
        ws: Worksheet to modify
        results (list): List of result dictionaries
        num_chromosomes (int): Number of chromosomes to display
        rotated (bool): If True, use rotated layout without absolute values
    """
    config = Config.get_instance()
    
    logger.debug(f"Creating grid layout for {len(results)} results with {num_chromosomes} chromosomes")
    logger.debug(f"Using {'rotated' if rotated else 'default'} layout")
    
    # Set layout labels based on rotation
    if rotated:
        col_labels = [str(i) for i in range(1, 13)]  # 1-12 down the side
        row_labels = list('ABCDEFGH')  # A-H across the top
    else:
        col_labels = DEFAULT_COL_LABELS  # A-H down the side
        row_labels = DEFAULT_ROW_LABELS  # 1-12 across the top
    
    # Set header row values
    setup_header_values(ws, row_labels, rotated)
    
    # Convert results list to a dictionary keyed by well ID
    result_dict = {r.get('well', ''): r for r in results if r.get('well') is not None}
    logger.debug(f"Created result dictionary with {len(result_dict)} valid wells")
    
    # Process each column division
    for col_idx, col_label in enumerate(col_labels):
        logger.debug(f"Processing {'row' if rotated else 'column'} {col_label}")
        # Starting row for this plate column
        start_row = col_idx * (num_chromosomes + 1) + 3
        
        # Add column label in column A
        col_label_cell = ws.cell(row=start_row, column=1)
        col_label_cell.value = col_label
        col_label_cell.font = Font(bold=True)
        col_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Process each row division
        for row_idx, row_label in enumerate(row_labels):
            # Generate well ID based on layout
            if rotated:
                well_id = f"{row_label}{col_label.zfill(2)}"
            else:
                well_id = f"{col_label}{row_label.zfill(2)}"
            
            # Determine well width based on layout
            well_width = 1 if rotated else 3
            
            # Starting column for this well
            start_col = row_idx * well_width + 2
            
            # Fill in data for this well without merging
            add_well_data(ws, start_row, start_col, well_id, result_dict.get(well_id), 
                         num_chromosomes, rotated)


def setup_header_values(ws, row_labels, rotated=False):
    """
    Set up the header row values without merging.
    
    Args:
        ws: Worksheet to modify
        row_labels (list): Labels for the horizontal divisions
        rotated (bool): If True, using rotated layout without absolute values
    """
    logger.debug(f"Setting up header values for {'rotated' if rotated else 'default'} layout")
    
    well_width = 1 if rotated else 3
    
    # Row 1: Horizontal headers
    for row_idx, row_label in enumerate(row_labels):
        cell_col = row_idx * well_width + 2
        cell = ws.cell(row=1, column=cell_col)
        cell.value = row_label
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        logger.debug(f"Set header {row_label} at column {cell_col}")
    
    # Row 2: Column labels based on layout
    for row_idx, row_label in enumerate(row_labels):
        base_col = row_idx * well_width + 2
        
        if rotated:
            # Single column: "rel."
            rel_cell = ws.cell(row=2, column=base_col)
            rel_cell.value = "rel."
            rel_cell.font = Font(size=9)
            rel_cell.alignment = Alignment(horizontal='center')
        else:
            # Three columns: blank, "abs.", "rel."
            first_cell = ws.cell(row=2, column=base_col)
            first_cell.value = ""
            
            abs_cell = ws.cell(row=2, column=base_col + 1)
            abs_cell.value = "abs."
            abs_cell.font = Font(size=9)
            abs_cell.alignment = Alignment(horizontal='center')
            
            rel_cell = ws.cell(row=2, column=base_col + 2)
            rel_cell.value = "rel."
            rel_cell.font = Font(size=9)
            rel_cell.alignment = Alignment(horizontal='center')
        
        logger.debug(f"Set subheaders for {row_label}")


def add_well_data(ws, start_row, start_col, well_id, result, num_chromosomes, rotated=False):
    """
    Add data for a single well without merging.
    
    Args:
        ws: Worksheet to modify
        start_row (int): Starting row for this well
        start_col (int): Starting column for this well
        well_id (str): Well identifier
        result (dict): Result dictionary for this well, or None if no data
        num_chromosomes (int): Number of chromosomes to display
        rotated (bool): If True, using rotated layout without absolute values
    """
    config = Config.get_instance()
    well_width = 1 if rotated else 3
    
    logger.debug(f"Adding data for well {well_id} at position ({start_row}, {start_col})")
    
    # Add the sample name to the first row
    name_cell = ws.cell(row=start_row, column=start_col)
    
    if result:
        sample_name = result.get('sample_name')
        if not sample_name:
            filename = result.get('filename', '')
            sample_name = os.path.splitext(filename)[0] if filename else well_id
            
        name_cell.value = sample_name
        name_cell.alignment = Alignment(horizontal='center', wrap_text=True)
        name_cell.font = Font(size=9)
        logger.debug(f"Added sample name: {sample_name}")
    else:
        name_cell.value = well_id
        name_cell.alignment = Alignment(horizontal='center')
        name_cell.font = Font(size=9, color='C0C0C0')
        logger.debug(f"No data for well, using well ID as placeholder")
    
    # Set blank values to cells that will be merged with the name cell
    for i in range(1, well_width):
        ws.cell(row=start_row, column=start_col + i).value = ""
    
    # Fill in chromosome data if we have results
    if result:
        _add_chromosome_data(ws, start_row, start_col, result, num_chromosomes, 
                           rotated, config)
    else:
        _add_empty_chromosome_labels(ws, start_row, start_col, num_chromosomes, 
                                   rotated, config)


def _add_chromosome_data(ws, start_row, start_col, result, num_chromosomes, rotated, config):
    """
    Add chromosome data for a well with results.
    
    Args:
        ws: Worksheet to modify
        start_row (int): Starting row for this well
        start_col (int): Starting column for this well
        result (dict): Result dictionary
        num_chromosomes (int): Number of chromosomes
        rotated (bool): Layout type
        config: Configuration instance
    """
    counts = result.get('counts', {})
    copy_numbers = result.get('copy_numbers', {})
    copy_number_states = result.get('copy_number_states', {})
    has_aneuploidy = result.get('has_aneuploidy', False)
    has_buffer_zone = result.get('has_buffer_zone', False)
    
    logger.debug(f"Well has aneuploidy: {has_aneuploidy}, buffer zone: {has_buffer_zone}")
    
    chromosome_keys = config.get_chromosome_keys()
    
    for chrom_idx, chrom_key in enumerate(chromosome_keys):
        chrom_row = start_row + 1 + chrom_idx
        
        if rotated:
            _add_rotated_chromosome_data(ws, chrom_row, start_col, chrom_key, 
                                       copy_numbers, copy_number_states, 
                                       has_aneuploidy, has_buffer_zone)
        else:
            _add_standard_chromosome_data(ws, chrom_row, start_col, chrom_key, 
                                        counts, copy_numbers, copy_number_states, 
                                        has_aneuploidy, has_buffer_zone)


def _add_rotated_chromosome_data(ws, chrom_row, start_col, chrom_key, copy_numbers, 
                               copy_number_states, has_aneuploidy, has_buffer_zone):
    """Add chromosome data for rotated layout (relative values only)."""
    rel_cell = ws.cell(row=chrom_row, column=start_col)
    rel_count = copy_numbers.get(chrom_key)
    
    if rel_count is not None:
        rel_cell.value = round(rel_count, 2)
        rel_cell.number_format = '0.00'
        rel_cell.font = Font(size=9)
        rel_cell.alignment = Alignment(horizontal='center')
    
    # Apply highlighting
    if has_buffer_zone:
        rel_cell.fill = PatternFill(start_color="B0B0B0", end_color="B0B0B0", fill_type="solid")
    elif has_aneuploidy:
        rel_cell.fill = ANEUPLOIDY_FILL
        
        if rel_count is not None:
            chrom_state = copy_number_states.get(chrom_key, 'euploid')
            if chrom_state == 'aneuploidy':
                rel_cell.fill = ANEUPLOIDY_VALUE_FILL


def _add_standard_chromosome_data(ws, chrom_row, start_col, chrom_key, counts, 
                                copy_numbers, copy_number_states, has_aneuploidy, 
                                has_buffer_zone):
    """Add chromosome data for standard layout (chromosome name, absolute, relative)."""
    chrom_label = f"Chr{chrom_key.replace('Chrom', '')}"
    
    # Chromosome name cell
    chrom_cell = ws.cell(row=chrom_row, column=start_col)
    chrom_cell.value = chrom_label
    chrom_cell.font = Font(size=9)
    
    # Absolute count cell
    abs_cell = ws.cell(row=chrom_row, column=start_col + 1)
    abs_count = counts.get(chrom_key, 0)
    abs_cell.value = abs_count if abs_count > 0 else None
    abs_cell.font = Font(size=9)
    abs_cell.alignment = Alignment(horizontal='center')
    
    # Relative count cell
    rel_cell = ws.cell(row=chrom_row, column=start_col + 2)
    rel_count = copy_numbers.get(chrom_key)
    if rel_count is not None:
        rel_cell.value = round(rel_count, 2)
        rel_cell.number_format = '0.00'
        rel_cell.font = Font(size=9)
        rel_cell.alignment = Alignment(horizontal='center')
    
    # Apply highlighting
    if has_buffer_zone:
        # Dark grey fill for buffer zone samples
        buffer_fill = PatternFill(start_color="B0B0B0", end_color="B0B0B0", fill_type="solid")
        chrom_cell.fill = buffer_fill
        abs_cell.fill = buffer_fill
        rel_cell.fill = buffer_fill
    elif has_aneuploidy:
        # Light purple fill for aneuploidy samples
        chrom_cell.fill = ANEUPLOIDY_FILL
        abs_cell.fill = ANEUPLOIDY_FILL
        rel_cell.fill = ANEUPLOIDY_FILL
        
        # Darker purple fill for individual aneuploidy chromosomes
        if rel_count is not None:
            chrom_state = copy_number_states.get(chrom_key, 'euploid')
            if chrom_state == 'aneuploidy':
                chrom_cell.fill = ANEUPLOIDY_VALUE_FILL
                abs_cell.fill = ANEUPLOIDY_VALUE_FILL
                rel_cell.fill = ANEUPLOIDY_VALUE_FILL


def _add_empty_chromosome_labels(ws, start_row, start_col, num_chromosomes, rotated, config):
    """Add chromosome labels for empty wells."""
    if not rotated:
        chromosome_keys = config.get_chromosome_keys()
        for chrom_idx, chrom_key in enumerate(chromosome_keys):
            chrom_label = f"Chr{chrom_key.replace('Chrom', '')}"
            chrom_row = start_row + 1 + chrom_idx
            chrom_cell = ws.cell(row=chrom_row, column=start_col)
            chrom_cell.value = chrom_label
            chrom_cell.font = Font(size=9, color='C0C0C0')


def apply_cell_merges(ws, num_chromosomes, rotated=False):
    """
    Apply all cell merges after setting all cell values.
    
    Args:
        ws: Worksheet to modify
        num_chromosomes (int): Number of chromosomes per well
        rotated (bool): If True, using rotated layout
    """
    logger.debug(f"Applying cell merges for {'rotated' if rotated else 'default'} layout")
    
    num_horizontal = 8 if rotated else 12
    well_width = 1 if rotated else 3
    well_height = num_chromosomes + 1
    num_vertical = 12 if rotated else 8
    
    # Merge cells in header row
    for row_idx in range(num_horizontal):
        cell_col = row_idx * well_width + 2
        ws.merge_cells(start_row=1, start_column=cell_col, 
                      end_row=1, end_column=cell_col + well_width - 1)
    
    # Merge vertical labels
    for col_idx in range(num_vertical):
        start_row = col_idx * well_height + 3
        ws.merge_cells(start_row=start_row, start_column=1, 
                      end_row=start_row + well_height - 1, end_column=1)
    
    # Merge well name cells
    for col_idx in range(num_vertical):
        for row_idx in range(num_horizontal):
            start_row = col_idx * well_height + 3
            start_col = row_idx * well_width + 2
            ws.merge_cells(start_row=start_row, start_column=start_col, 
                          end_row=start_row, end_column=start_col + well_width - 1)


def apply_formatting(ws, num_chromosomes, rotated=False):
    """
    Apply formatting to the entire worksheet with proper borders.
    
    Args:
        ws: Worksheet to modify
        num_chromosomes (int): Number of chromosomes per well
        rotated (bool): If True, using rotated layout
    """
    logger.debug(f"Applying formatting for {'rotated' if rotated else 'default'} layout")
    
    well_height = num_chromosomes + 1
    num_vertical = 12 if rotated else 8
    num_horizontal = 8 if rotated else 12
    well_width = 1 if rotated else 3
    
    max_row = num_vertical * well_height + 2
    max_col = num_horizontal * well_width + 1
    
    logger.debug(f"Worksheet dimensions: {max_row}x{max_col}")
    
    # Apply thin borders to all cells
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    
    # Apply thick borders for plate boundaries
    apply_plate_boundaries(ws, num_chromosomes, rotated)
    
    # Set column widths
    ws.column_dimensions['A'].width = 3
    
    for i in range(1, max_col):
        col_letter = get_column_letter(i + 1)
        if rotated:
            ws.column_dimensions[col_letter].width = 6
        else:
            width = 5 if i % well_width == 1 else 6
            ws.column_dimensions[col_letter].width = width
    
    # Freeze the header rows and first column
    ws.freeze_panes = ws.cell(row=3, column=2)


def apply_plate_boundaries(ws, num_chromosomes, rotated=False):
    """
    Apply thick borders around each well and at the plate boundaries.
    
    Args:
        ws: Worksheet to modify
        num_chromosomes (int): Number of chromosomes per well
        rotated (bool): If True, using rotated layout
    """
    logger.debug(f"Applying plate boundaries for {'rotated' if rotated else 'default'} layout")
    
    well_height = num_chromosomes + 1
    num_vertical = 12 if rotated else 8
    num_horizontal = 8 if rotated else 12
    well_width = 1 if rotated else 3
    
    # Apply thick borders around each well
    for col_idx in range(num_vertical):
        start_row = col_idx * well_height + 3
        end_row = start_row + well_height - 1
        
        for row_idx in range(num_horizontal):
            start_col = row_idx * well_width + 2
            end_col = start_col + well_width - 1
            
            # Apply borders to each cell in this well
            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    cell = ws.cell(row=r, column=c)
                    
                    # Determine border sides
                    left_side = thick if c == start_col else thin
                    right_side = thick if c == end_col else thin
                    top_side = thick if r == start_row else thin
                    bottom_side = thick if r == end_row else thin
                    
                    cell.border = Border(
                        left=left_side, right=right_side,
                        top=top_side, bottom=bottom_side
                    )