#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
List report generation module for ddQuint with tabular format and buffer zone support.

Generates Excel reports in list format with chromosome data organized in columns.
Includes copy number highlighting for aneuploidies and buffer zones with proper
column-first well sorting.
"""

import os
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

from ..config import Config, ReportGenerationError

logger = logging.getLogger(__name__)


def create_list_report(results, output_path):
    """
    Create a list-format Excel report with chromosome data in columns.
    
    Args:
        results (list): List of result dictionaries for each well
        output_path (str): Path to save the Excel report
        
    Returns:
        str: Path to the saved Excel report
        
    Raises:
        ReportGenerationError: If report creation fails
        ValueError: If results data is invalid
    """
    config = Config.get_instance()
    
    if not results:
        error_msg = "No results provided for list report generation"
        logger.error(error_msg)
        raise ValueError(error_msg)
    
    # Get the number of chromosomes from config
    chromosome_keys = config.get_chromosome_keys()
    num_chromosomes = len(chromosome_keys)
    
    logger.debug(f"Creating list report for {len(results)} results with {num_chromosomes} chromosomes")
    logger.debug(f"Output path: {output_path}")
    
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "List Results"
        logger.debug("Created new workbook")
        
        # Sort results by well ID in column-first order
        sorted_results = sorted(results, key=lambda x: parse_well_id_column_first(x.get('well', '')))
        
        # Set up headers with proper structure
        setup_headers(ws, chromosome_keys)
        
        # Fill in data for each well
        fill_well_data(ws, sorted_results, chromosome_keys, config)
        
        # Apply formatting
        apply_formatting(ws, len(sorted_results), chromosome_keys)
        
        # Save the workbook
        wb.save(output_path)
        logger.debug(f"List report saved successfully to {output_path}")
        return output_path
        
    except Exception as e:
        error_msg = f"Error creating list report for {os.path.basename(output_path)}: {str(e)}"
        logger.error(error_msg)
        logger.debug(f"Error details: {str(e)}", exc_info=True)
        raise ReportGenerationError(error_msg) from e


def parse_well_id_column_first(well_id):
    """
    Parse well ID to support column-first sorting.
    
    Args:
        well_id (str): Well identifier like 'A01', 'B12', etc.
        
    Returns:
        tuple: (column_number, row_number) for column-first sorting
    """
    if not well_id:
        return (999, 999)  # Put empty well IDs at the end
    
    # Extract letter(s) for row and number(s) for column
    row_part = ''
    col_part = ''
    
    for char in well_id:
        if char.isalpha():
            row_part += char
        elif char.isdigit():
            col_part += char
    
    # Convert row letters to number (A=1, B=2, etc.)
    if row_part:
        row_number = 0
        for i, char in enumerate(reversed(row_part.upper())):
            row_number += (ord(char) - ord('A') + 1) * (26 ** i)
    else:
        row_number = 999  # Put malformed wells at the end
    
    # Convert column to integer
    try:
        col_number = int(col_part) if col_part else 0
    except ValueError:
        col_number = 999
    
    # Return (column, row) for column-first sorting
    return (col_number, row_number)


def setup_headers(ws, chromosome_keys):
    """
    Set up headers with proper merging and structure.
    
    Args:
        ws: Worksheet to modify
        chromosome_keys (list): List of chromosome identifiers
    """
    num_chromosomes = len(chromosome_keys)
    
    # Row 1: Main headers
    ws.cell(row=1, column=1, value="Well")
    ws.cell(row=2, column=1, value="")
    
    ws.cell(row=1, column=2, value="Sample")
    ws.cell(row=2, column=2, value="")
    
    # Relative Copy Number section
    rel_start = 3
    rel_end = 2 + num_chromosomes
    ws.cell(row=1, column=rel_start, value="Relative Copy Number")
    for i in range(rel_start, rel_end + 1):
        ws.cell(row=1, column=i + 1, value="")
    
    # Absolute Copy Number section
    abs_start = rel_end + 1
    abs_end = abs_start + num_chromosomes - 1
    ws.cell(row=1, column=abs_start, value="Absolute Copy Number")
    for i in range(abs_start, abs_end + 1):
        ws.cell(row=1, column=i + 1, value="")
    
    # Merge cells for headers
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=rel_start, end_row=1, end_column=rel_end)
    ws.merge_cells(start_row=1, start_column=abs_start, end_row=1, end_column=abs_end)
    
    # Row 2: Chromosome headers
    for i, chrom_key in enumerate(chromosome_keys):
        chrom_label = f"Chr{chrom_key.replace('Chrom', '')}"
        ws.cell(row=2, column=rel_start + i, value=chrom_label)
        ws.cell(row=2, column=abs_start + i, value=chrom_label)
    
    # Apply formatting to headers
    for cell in ws[1]:
        if cell.value:
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[2]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')


def fill_well_data(ws, sorted_results, chromosome_keys, config):
    """
    Fill in data for each well with buffer zone and aneuploidy highlighting.
    
    Args:
        ws: Worksheet to modify
        sorted_results (list): Sorted list of result dictionaries
        chromosome_keys (list): List of chromosome identifiers
        config: Configuration instance
    """
    num_chromosomes = len(chromosome_keys)
    rel_start = 3
    abs_start = rel_start + num_chromosomes
    
    for row_idx, result in enumerate(sorted_results, start=3):
        well_id = result.get('well', '')
        
        # Well ID
        well_cell = ws.cell(row=row_idx, column=1, value=well_id)
        well_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sample name
        sample_name = result.get('sample_name', '')
        if not sample_name:
            filename = result.get('filename', '')
            sample_name = os.path.splitext(filename)[0] if filename else well_id
        
        sample_cell = ws.cell(row=row_idx, column=2, value=sample_name)
        sample_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Get data for this well
        counts = result.get('counts', {})
        copy_numbers = result.get('copy_numbers', {})
        copy_number_states = result.get('copy_number_states', {})
        has_aneuploidy = result.get('has_aneuploidy', False)
        has_buffer_zone = result.get('has_buffer_zone', False)
        
        # Determine row-level highlighting - buffer zone trumps aneuploidy
        row_fill = None
        if has_buffer_zone:
            row_fill = PatternFill(start_color="B0B0B0", end_color="B0B0B0", fill_type="solid")
        elif has_aneuploidy:
            row_fill = PatternFill(start_color="E6B8E6", end_color="E6B8E6", fill_type="solid")
        
        # Apply highlighting to Well and Sample cells
        if row_fill:
            well_cell.fill = row_fill
            sample_cell.fill = row_fill
        
        # Fill in relative and absolute copy numbers
        _fill_chromosome_data(ws, row_idx, rel_start, abs_start, chromosome_keys, 
                             copy_numbers, counts, copy_number_states, row_fill, 
                             has_buffer_zone, has_aneuploidy)


def _fill_chromosome_data(ws, row_idx, rel_start, abs_start, chromosome_keys, 
                         copy_numbers, counts, copy_number_states, row_fill, 
                         has_buffer_zone, has_aneuploidy):
    """
    Fill chromosome data with appropriate highlighting.
    
    Args:
        ws: Worksheet to modify
        row_idx (int): Current row index
        rel_start (int): Starting column for relative data
        abs_start (int): Starting column for absolute data
        chromosome_keys (list): List of chromosome identifiers
        copy_numbers (dict): Copy number values
        counts (dict): Absolute count values
        copy_number_states (dict): Copy number state classifications
        row_fill: PatternFill for row-level highlighting
        has_buffer_zone (bool): Whether sample has buffer zone
        has_aneuploidy (bool): Whether sample has aneuploidy
    """
    for i, chrom_key in enumerate(chromosome_keys):
        # Relative copy numbers
        rel_cell = ws.cell(row=row_idx, column=rel_start + i)
        rel_count = copy_numbers.get(chrom_key)
        if rel_count is not None:
            rel_cell.value = round(rel_count, 2)
            rel_cell.number_format = '0.00'
        else:
            rel_cell.value = ""
        rel_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Absolute copy numbers
        abs_cell = ws.cell(row=row_idx, column=abs_start + i)
        abs_count = counts.get(chrom_key, 0)
        abs_cell.value = abs_count if abs_count > 0 else ""
        abs_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply highlighting
        if has_buffer_zone:
            # Buffer zone samples get uniform dark grey fill
            rel_cell.fill = row_fill
            abs_cell.fill = row_fill
        elif has_aneuploidy:
            chrom_state = copy_number_states.get(chrom_key, 'euploid')
            if chrom_state == 'aneuploidy':
                # Individual chromosome aneuploidy highlighting
                chrom_fill = PatternFill(start_color="D070D0", end_color="D070D0", fill_type="solid")
                rel_cell.fill = chrom_fill
                abs_cell.fill = chrom_fill
            else:
                # Non-aneuploidy chromosome in aneuploidy sample
                rel_cell.fill = row_fill
                abs_cell.fill = row_fill


def apply_formatting(ws, num_results, chromosome_keys):
    """
    Apply borders, column widths, and freeze panes.
    
    Args:
        ws: Worksheet to modify
        num_results (int): Number of result rows
        chromosome_keys (list): List of chromosome identifiers
    """
    num_chromosomes = len(chromosome_keys)
    rel_start = 3
    abs_start = rel_start + num_chromosomes
    max_col = abs_start + num_chromosomes - 1
    max_row = num_results + 2
    
    # Border styles
    thick = Side(style='thick')
    
    # Apply borders
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            
            # Initialize borders
            left_border = None
            right_border = None
            top_border = None
            bottom_border = None
            
            # Thick borders for table outline
            if row == 1:
                top_border = thick
            if row == max_row:
                bottom_border = thick
            if col == 1:
                left_border = thick
            if col == max_col:
                right_border = thick
            
            # Thick borders around sections
            if col == rel_start and row >= 1:
                left_border = thick
            if col == rel_start + num_chromosomes - 1 and row >= 1:
                right_border = thick
                
            # Thick border at bottom of header row
            if row == 2:
                bottom_border = thick
                
            # Apply border if at least one side has a style
            if any([left_border, right_border, top_border, bottom_border]):
                cell.border = Border(left=left_border, right=right_border, 
                                   top=top_border, bottom=bottom_border)
    
    # Set column widths
    ws.column_dimensions['A'].width = 8  # Well
    ws.column_dimensions['B'].width = 15  # Sample
    
    # Set widths for chromosome columns
    for col in range(rel_start, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 8
    
    # Freeze panes (freeze top 2 rows and first 2 columns)
    ws.freeze_panes = ws.cell(row=3, column=3)